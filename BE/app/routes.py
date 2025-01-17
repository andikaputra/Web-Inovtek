from flask import Blueprint, request,send_from_directory, abort,send_file, jsonify, current_app
from .models import db, SoalJenis, Waktu, QuizKode, Soal, SoalJawaban, Peserta, PesertaNilai, PesertaJawaban, Users, SesiMainVR, SesiMainAR,LogSession,LogSessionKota
from flask_cors import CORS
from sqlalchemy.exc import IntegrityError  # Import IntegrityError
from sqlalchemy.orm import joinedload
from sqlalchemy import func, desc, literal
from sqlalchemy import desc
from datetime import datetime
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import os
import json 
import openpyxl 
from docx import Document

 
quiz_bp = Blueprint('quiz', __name__)

CORS(quiz_bp)

@quiz_bp.route('/', methods=['GET'])
def index():
    return jsonify({"message": "Welcome to the Quiz API"}), 200

# =================================QUIZ CODE==========================================================
# CREATE
@quiz_bp.route('/quiz_kode', methods=['POST'])
def create_quiz_kode():
    data = request.get_json()
    new_quiz_kode = QuizKode(
        kode=data['kode'],
        created_at=datetime.utcnow()
    )
    db.session.add(new_quiz_kode)
    db.session.commit()
    return jsonify({'message': 'QuizKode created successfully', 'data': data}), 201

# READ (GET ALL)
@quiz_bp.route('/quiz_kode', methods=['GET'])
def get_quiz_kode():
    quiz_kodes = QuizKode.query.filter_by(deleted_at=None).all()
    output = []
    for quiz in quiz_kodes:
        quiz_data = {'id': quiz.id, 'kode': quiz.kode, 'mulai': quiz.mulai, 'created_at': quiz.created_at, 'updated_at': quiz.updated_at}
        output.append(quiz_data)
    return jsonify({'data': output}), 200

# READ (GET BY ID)
@quiz_bp.route('/quiz_kode/<int:id>', methods=['GET'])
def get_quiz_kode_by_id(id):
    quiz = QuizKode.query.filter_by(id=id, deleted_at=None).first()
    if not quiz:
        return jsonify({'message': 'QuizKode not found'}), 404
    quiz_data = {'id': quiz.id, 'kode': quiz.kode, 'created_at': quiz.created_at, 'updated_at': quiz.updated_at}
    return jsonify({'data': quiz_data}), 200

# READ (GET BY Kode)
@quiz_bp.route('/quiz/kode', methods=['GET'])
def get_quiz_kode_by_kode():
    kode = request.args.get('kode')
    
    if not kode:
        return jsonify({'error': 'Kode tidak disediakan'}), 400

    quiz = QuizKode.query.filter_by(kode=kode, mulai=True, deleted_at=None).first()
    if not quiz:
        return jsonify({'message': 'QuizKode not found'}), 404
    quiz_data = {'id': quiz.id, 'kode': quiz.kode, 'created_at': quiz.created_at, 'updated_at': quiz.updated_at}
    return jsonify({'data': quiz_data}), 200

# UPDATE
@quiz_bp.route('/quiz_kode/<int:id>', methods=['PUT'])
def update_quiz_kode(id):
    data = request.get_json()
    quiz = QuizKode.query.filter_by(id=id, deleted_at=None).first()
    if not quiz:
        return jsonify({'message': 'QuizKode not found'}), 404
    quiz.kode = data.get('kode', quiz.kode)
    quiz.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'message': 'QuizKode updated successfully'}), 200

# DELETE
@quiz_bp.route('/quiz_kode/<int:id>', methods=['DELETE'])
def delete_quiz_kode(id):
    quiz = QuizKode.query.filter_by(id=id, deleted_at=None).first()
    if not quiz:
        return jsonify({'message': 'QuizKode not found'}), 404
    quiz.deleted_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'message': 'QuizKode deleted successfully'}), 200

@quiz_bp.route('/quiz_mulai/<int:quiz_id>', methods=['PUT'])
def update_mulai(quiz_id):
    try:
        # Cari QuizKode berdasarkan ID
        quiz = QuizKode.query.filter_by(id=quiz_id, deleted_at=None).first()
        if not quiz:
            return jsonify({'message': 'QuizKode not found'}), 404

        # Update field mulai menjadi True
        quiz.mulai = True
        quiz.updated_at = datetime.utcnow()

        # Simpan perubahan ke database
        db.session.commit()

        return jsonify({'message': 'Quiz updated successfully', 'quiz': {
            'id': quiz.id,
            'kode': quiz.kode,
            'mulai': quiz.mulai,
            'updated_at': quiz.updated_at
        }}), 200

    except Exception as e:
        return jsonify({'message': 'An error occurred', 'error': str(e)}), 500


# =====================================SOAL JENIS==========================================================

@quiz_bp.route('/soal_jenis', methods=['GET'])
def get_soal_jenis():
    # Mendapatkan semua data dari tabel SoalJenis
    soalJenis = SoalJenis.query.all()
    # Memformat data dalam bentuk JSON
    data = [{"id": sj.id, "nama": sj.nama} for sj in soalJenis]
    return jsonify({"data": data}), 200

# =====================================WAKTU===========================================================
@quiz_bp.route('/waktu', methods=['GET'])
def get_waktu():
    # Mendapatkan semua data dari tabel SoalJenis
    waktu = Waktu.query.all()
    # Memformat data dalam bentuk JSON
    data = [{"id": wk.id, "nama": wk.nama, "detik": wk.detik} for wk in waktu]
    return jsonify({"data": data}), 200

# ======================================SOAL==========================================================


# Tentukan folder penyimpanan file
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}
  
def allowed_file(filename):
    """Memeriksa apakah file memiliki ekstensi yang diizinkan."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@quiz_bp.route('/soal', methods=['POST'])
def upload_file():
    # Mengambil data dari form request
    id_quiz_kode = request.form['id_quiz_kode']
    id_soal_jenis = request.form['id_soal_jenis']
    id_waktu = request.form['id_waktu']
    pertanyaan = request.form['pertanyaan']
    layout = request.form['layout']
    file_path = None
    jenis_file = None

    # Memeriksa apakah file ada dalam request
    if 'file' in request.files:
        file = request.files['file']
        filename = secure_filename(file.filename)
        
        # Gunakan current_app untuk mengakses konfigurasi app
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Menentukan jenis file dari nama file
        jenis_file = filename.rsplit('.', 1)[1].lower()
    
    # Menyimpan data ke database
    new_soal = Soal(
        id_quiz_kode=id_quiz_kode,
        id_soal_jenis=id_soal_jenis,
        id_waktu=id_waktu,
        pertanyaan=pertanyaan,
        file=file_path,
        jenis_file=jenis_file,
        layout=layout,
        created_at=datetime.utcnow()
    )
    
    db.session.add(new_soal)
    db.session.commit()

    # Ambil data dari answersQuiz
    index = 0

    # Mengambil semua data dari form
    while True:
        answer_data = request.form.get(f'answers[{index}]')
        if answer_data is None:
            break  # Keluar dari loop jika tidak ada data untuk indeks ini

        # Konversi JSON string ke Python dictionary
        answer_object = json.loads(answer_data)

        # Cek dan simpan file jawaban jika ada
        file_jawaban_path = None 
        if f'file_{index}' in request.files:
            file_jawaban = request.files[f'file_{index}']
            filename_jawaban = secure_filename(file_jawaban.filename)
            
            # Path untuk menyimpan file jawaban
            file_jawaban_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename_jawaban)
            file_jawaban.save(file_jawaban_path)
            
            # Simpan path file ke text_jawaban jika ada file
            text_jawaban = file_jawaban_path
        else:
            # Jika tidak ada file, simpan teks biasa dari `answer_object['text']`
            text_jawaban = answer_object['text']

        # Membuat entri jawaban di database
        new_jawaban = SoalJawaban(
            id_soal=new_soal.id,
            text_jawaban=text_jawaban,
            correct=answer_object['correct'],
            bobot=1 if answer_object['correct'] else 0
        )
        
        db.session.add(new_jawaban)
        db.session.commit()
        index += 1

    return jsonify({'success': 'File uploaded successfully'}), 200


@quiz_bp.route('/soal/<int:id>', methods=['DELETE'])
def delete_soal(id):
    # Mencari soal berdasarkan ID
    soal = Soal.query.get(id)
    
    # Jika soal tidak ditemukan
    if not soal:
        return jsonify({'error': 'Soal not found'}), 404
    
    # Hapus semua jawaban yang terkait dengan soal ini
    SoalJawaban.query.filter_by(id_soal=id).delete()

    # Menghapus soal dari database
    db.session.delete(soal)
    db.session.commit()
    
    return jsonify({'message': 'Soal deleted successfully'}), 200

@quiz_bp.route('/soal/<int:id_quiz_kode>', methods=['GET'])
def get_soal_by_quiz_kode(id_quiz_kode):
    # Mendapatkan semua soal berdasarkan id_quiz_kode
    soals = Soal.query.filter_by(id_quiz_kode=id_quiz_kode).all()
    
    # Jika tidak ada soal ditemukan
    if not soals:
        return jsonify({'error': 'No questions found for the specified quiz code'}), 404
    
    # Membuat respons dengan struktur JSON yang mencakup soal dan jawabannya
    soal_data = []
    for soal in soals:
        jawaban_data = []
        for jawaban in soal.soal_jawaban:
            jawaban_data.append({
                'id': jawaban.id,
                'text_jawaban': jawaban.text_jawaban,
                'correct': jawaban.correct,
                'bobot': jawaban.bobot,
                'created_at': jawaban.created_at,
                'updated_at': jawaban.updated_at,
                'deleted_at': jawaban.deleted_at
            })
        
        soal_data.append({
            'id': soal.id,
            'id_quiz_kode': soal.id_quiz_kode,
            'id_soal_jenis': soal.id_soal_jenis,
            'id_waktu': soal.id_waktu,
            'pertanyaan': soal.pertanyaan,
            'file': soal.file,
            'jenis_file': soal.jenis_file,
            'layout': soal.layout,
            'created_at': soal.created_at,
            'updated_at': soal.updated_at,
            'deleted_at': soal.deleted_at,
            'jawaban': jawaban_data  # Menyertakan data jawaban terkait
        })
    
    return jsonify(soal_data), 200
 

@quiz_bp.route('/soal/ujian/<int:id_quiz_kode>/<int:offset>', methods=['GET'])
def get_soal_ujian_by_quiz_kode(id_quiz_kode, offset=0):
    # Query untuk mengambil soal dan waktu terkait
    soal = Soal.query.filter_by(
        id_quiz_kode=id_quiz_kode,
        deleted_at=None  # Memastikan hanya data yang belum dihapus
    ).options(
        joinedload(Soal.soal_jawaban),  # Memuat soal_jawaban
        joinedload(Soal.waktu)  # Memuat waktu yang berelasi dengan soal
    ).limit(1).offset(offset).first()

    # Jika data soal ditemukan, kembalikan dalam format JSON
    if soal:
        return jsonify({
            "id": soal.id,
            "id_quiz_kode": soal.id_quiz_kode,
            "id_soal_jenis": soal.id_soal_jenis,
            "id_waktu": soal.id_waktu,
            "waktu": {
                "id": soal.waktu.id,
                "nama": soal.waktu.nama,
                "detik": soal.waktu.detik
            } if soal.waktu else None,
            "pertanyaan": soal.pertanyaan,
            "file": soal.file,
            "jenis_file": soal.jenis_file,
            "layout": soal.layout,
            "created_at": soal.created_at,
            "updated_at": soal.updated_at,
            "deleted_at": soal.deleted_at,
            "soal_jawaban": [
                {
                    "id": jawaban.id,
                    "text_jawaban": jawaban.text_jawaban,
                    "correct": jawaban.correct,
                    "bobot": jawaban.bobot
                }
                for jawaban in soal.soal_jawaban
            ]
        })
    else:
        return jsonify({"error": "Soal not found or deleted"}), 404


@quiz_bp.route('/copy_soal', methods=['POST'])
def copy_soal():
    try:
        # Ambil data dari request
        data = request.get_json()
        source_quiz_id = data.get('source_quiz_id')
        target_quiz_kode = data.get('target_quiz_kode')  # Gunakan kode untuk target quiz

        if not source_quiz_id or not target_quiz_kode:
            return jsonify({"error": "source_quiz_id and target_quiz_kode are required"}), 400

        # Pastikan source_quiz_id valid
        source_quiz = QuizKode.query.get(source_quiz_id)

        if not source_quiz:
            return jsonify({"error": "Source quiz not found"}), 404

        # Periksa apakah target_quiz_kode sudah ada
        existing_target_quiz = QuizKode.query.filter_by(kode=target_quiz_kode, deleted_at=None).first()
        if existing_target_quiz:
            return jsonify({"error": f"Quiz code '{target_quiz_kode}' already exists in the database"}), 400

        # Buat entri baru untuk target_quiz
        new_target_quiz = QuizKode(
            kode=target_quiz_kode,
            mulai=False,
            created_at=datetime.utcnow()
        )
        db.session.add(new_target_quiz)
        db.session.flush()  # Flush untuk mendapatkan ID target_quiz yang baru dibuat

        # Salin soal dari source_quiz ke target_quiz
        source_soal_list = Soal.query.filter_by(id_quiz_kode=source_quiz_id, deleted_at=None).all()
        for soal in source_soal_list:
            # Buat salinan soal baru
            new_soal = Soal(
                id_quiz_kode=new_target_quiz.id,  # Gunakan ID target_quiz yang baru dibuat
                id_soal_jenis=soal.id_soal_jenis,
                id_waktu=soal.id_waktu,
                pertanyaan=soal.pertanyaan,
                file=soal.file,
                jenis_file=soal.jenis_file,
                layout=soal.layout,
                created_at=datetime.utcnow()
            )
            db.session.add(new_soal)
            db.session.flush()  # Dapatkan ID soal baru sebelum commit

            # Salin jawaban untuk soal ini
            source_jawaban_list = SoalJawaban.query.filter_by(id_soal=soal.id, deleted_at=None).all()
            for jawaban in source_jawaban_list:
                new_jawaban = SoalJawaban(
                    id_soal=new_soal.id,
                    text_jawaban=jawaban.text_jawaban,
                    correct=jawaban.correct,
                    bobot=jawaban.bobot,
                    created_at=datetime.utcnow()
                )
                db.session.add(new_jawaban)

        # Commit semua perubahan
        db.session.commit()
        return jsonify({
            "message": "Soal and answers successfully copied",
            "target_quiz_id": new_target_quiz.id,
            "target_quiz_kode": new_target_quiz.kode
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@quiz_bp.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    # Cek apakah file ada di folder yang ditentukan
    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(file_path):
        abort(404, description="File not found")

    # Kirim file ke klien
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

# Example for inline view without download
@quiz_bp.route('/view/<filename>', methods=['GET'])
def view_file(filename):
    UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'upload')
    current_app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    if not os.path.exists(file_path):
        abort(404, description="File not found")

    return send_file(file_path)

@quiz_bp.route('/upload-soal-excel', methods=['POST'])
def upload_soal():
    try:
        # Get `id_quiz_kode` from request form
        id_quiz_kode = request.form.get('id_quiz_kode')
        if not id_quiz_kode:
            return jsonify({"error": "id_quiz_kode is required"}), 400

        # Check if a file is included in the request
        if 'file' not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400

        # Save file to server
        filename = secure_filename(file.filename)
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        # Open the uploaded Excel file
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Iterate over rows and process data
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming first row is header
            pertanyaan, jawaban_data = row  # Remove `id_quiz_kode` from Excel row

            # Insert into Soal table
            soal = Soal(
                id_quiz_kode=id_quiz_kode,  # Use `id_quiz_kode` from request
                id_soal_jenis=1,
                id_waktu=1,
                pertanyaan=pertanyaan,
                layout=1,
                created_at=datetime.utcnow(),
            )
            db.session.add(soal)
            db.session.flush()  # Get the inserted ID for the soal

            # Process jawaban_data (assuming it is a JSON-like string)
            if jawaban_data:
                jawaban_list = eval(jawaban_data)  # Convert string to list of dictionaries
                for jawaban in jawaban_list:
                    soal_jawaban = SoalJawaban(
                        id_soal=soal.id,
                        text_jawaban=jawaban.get('text_jawaban'),
                        correct=jawaban.get('correct'),
                        bobot=jawaban.get('bobot'),
                        created_at=datetime.utcnow(),
                    )
                    db.session.add(soal_jawaban)

        db.session.commit()
        return jsonify({"message": "Data successfully uploaded and processed"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
    # finally:
        # os.remove(file_path)  # Remove file after processing

# API to upload Word file and process data
@quiz_bp.route('/upload-soal-docs', methods=['POST'])
def upload_soal_docs():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    # Save the file
    filename = secure_filename(file.filename)
    file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    file.save(file_path)

    try:
        # Get `id_quiz_kode` from request form
        id_quiz_kode = request.form.get('id_quiz_kode')
        if not id_quiz_kode:
            return jsonify({"error": "id_quiz_kode is required"}), 400

        # Read the Word document
        document = Document(file_path)
        current_soal = None

        for paragraph in document.paragraphs:
            text = paragraph.text.strip()

            # Skip empty lines
            if not text:
                continue

            # Detect question (Soal)
            if text.startswith("Q:"):
                pertanyaan = text[2:].strip()
                current_soal = Soal(
                    id_quiz_kode=id_quiz_kode,  # Example value, replace with actual quiz ID
                    id_soal_jenis=1,  # Example value, replace with actual jenis ID
                    id_waktu=1,  # Example value, replace with actual waktu ID
                    pertanyaan=pertanyaan,
                    layout=1,  # Example value
                    created_at=datetime.utcnow(),
                )
                db.session.add(current_soal)
                db.session.flush()  # Get the inserted ID
            elif text.startswith("A:") and current_soal:
                # Detect answer (SoalJawaban)
                answer_data = text[2:].strip().split(";")
                text_jawaban = answer_data[0]
                correct = answer_data[1].lower() == "true"
                bobot = int(answer_data[2]) if len(answer_data) > 2 else 0

                soal_jawaban = SoalJawaban(
                    id_soal=current_soal.id,
                    text_jawaban=text_jawaban,
                    correct=correct,
                    bobot=bobot,
                    created_at=datetime.utcnow(),
                )
                db.session.add(soal_jawaban)

        db.session.commit()
        return jsonify({"message": "Data successfully uploaded and processed"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

 
 # ===============================PESERTA==============================================
@quiz_bp.route('/peserta', methods=['GET'])
def get_peserta():
    pes = Peserta.query.filter_by(deleted_at=None).all()
    output = []
    for peserta in pes:
        peserta_data = {
            'id': peserta.id,
            'id_quiz_kode': peserta.id_quiz_kode,
            'kode_unik': peserta.kode_unik,
            'created_at': peserta.created_at,
            'updated_at': peserta.updated_at,
            'deleted_at': peserta.deleted_at,
            # tambahkan field lain sesuai model Anda
        }
        output.append(peserta_data)
    return jsonify({'data': output}), 200


@quiz_bp.route('/peserta/<int:id_quiz_kode>', methods=['GET'])
def get_peserta_by_quiz_kode(id_quiz_kode):
    try:
        # Query to get all participants with the specified id_quiz_kode
        peserta_list = Peserta.query.filter_by(id_quiz_kode=id_quiz_kode, deleted_at=None).all()
        
        # Check if any participants are found
        if not peserta_list:
            return jsonify({'message': 'No participants found for the specified quiz code'}), 404

        # Prepare response data
        peserta_data = []
        for peserta in peserta_list:
            peserta_data.append({
                'id': peserta.id,
                'id_quiz_kode': peserta.id_quiz_kode,
                'kode_unik': peserta.kode_unik,
                'created_at': peserta.created_at,
                'updated_at': peserta.updated_at,
            })

        return jsonify(peserta_data), 200

    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': 'An error occurred while retrieving participants'}), 500


@quiz_bp.route('/peserta', methods=['POST'])
def create_peserta():
    data = request.get_json()

    # Cek jika peserta dengan `id_quiz_kode` dan `kode_unik` sudah ada, tetapi dihapus
    pes = Peserta.query.filter_by(
        id_quiz_kode=data['id_quiz_kode'],
        kode_unik=data['kode_unik'], 
        deleted_at=None
    ).first()
    if pes:
        pes.deleted_at = datetime.utcnow()
        db.session.commit()

    # Buat peserta baru
    new_peserta = Peserta(
        id_quiz_kode=data['id_quiz_kode'],
        kode_unik=data['kode_unik'],
        created_at=datetime.utcnow()
    )
    db.session.add(new_peserta)
    db.session.commit()

    # Ambil data lengkap peserta yang baru saja dibuat untuk ditampilkan
    peserta_data = {
        'id': new_peserta.id,
        'id_quiz_kode': new_peserta.id_quiz_kode,
        'kode_unik': new_peserta.kode_unik,
        'created_at': new_peserta.created_at,
        'deleted_at': new_peserta.deleted_at
    }

    return jsonify({'message': 'Peserta created successfully', 'data': peserta_data}), 200


@quiz_bp.route('/delete_peserta', methods=['DELETE'])
def delete_peserta_by_id_quiz_kode():
    try:
        # Ambil id_quiz_kode dari query parameter
        id_quiz_kode = request.args.get('id_quiz_kode')
        if not id_quiz_kode:
            return jsonify({'message': 'id_quiz_kode is required'}), 400

        # Query semua data dengan id_quiz_kode yang sesuai dan belum di-soft delete
        peserta_to_soft_delete = Peserta.query.filter_by(id_quiz_kode=id_quiz_kode, deleted_at=None).all()
        if not peserta_to_soft_delete:
            return jsonify({'message': f'No Peserta found with id_quiz_kode={id_quiz_kode}'}), 404

        # Soft delete semua data dengan mengisi `deleted_at`
        for peserta in peserta_to_soft_delete:
            peserta.deleted_at = datetime.utcnow()

        db.session.commit()

        return jsonify({'message': f'Successfully soft deleted Peserta with id_quiz_kode={id_quiz_kode}'}), 200
    except Exception as e:
        return jsonify({'message': 'An error occurred', 'error': str(e)}), 500


    # ===============================PESERTA NILAI=============================================PESERTA
@quiz_bp.route('/peserta_nilai/<int:id_peserta>/<int:id_quiz_kode>', methods=['GET'])
def get_peserta_nilai(id_peserta, id_quiz_kode):
    # Query untuk mengambil data peserta_nilai berdasarkan id_peserta, id_quiz_kode, dan deleted_at NULL
    data = PesertaNilai.query.filter_by(
        id_peserta=id_peserta,
        id_quiz_kode=id_quiz_kode,
        deleted_at=None  # Hanya mengambil data yang belum dihapus
    ).first()

    # Jika tidak ditemukan, kembalikan pesan error
    if not data:
        return jsonify({'message': 'Data not found'}), 404

    # Mengonversi data ke dalam format yang bisa dikembalikan sebagai JSON
    nilai_data = {
        'id': data.id,
        'id_quiz_kode': data.id_quiz_kode,
        'id_peserta': data.id_peserta,
        'nilai': data.nilai,  
        'created_at': data.created_at,
        'updated_at': data.updated_at,
        'deleted_at': data.deleted_at
    }

    return jsonify({'message': 'Data retrieved successfully', 'data': nilai_data}), 200

 
@quiz_bp.route('/submit_answers', methods=['POST'])
def submit_answers():
    data = request.get_json()
    id_quiz_kode = data.get('id_quiz_kode')
    id_peserta = data.get('id_peserta')
    answers = data.get('answers', [])

    if not id_quiz_kode or not id_peserta or not answers:
        return jsonify({'error': 'Data tidak lengkap'}), 400

    total_score = 0

    # Loop melalui setiap jawaban
    for answer in answers:
        id_soal = answer.get('id_soal')
        jawaban_text = answer.get('jawaban')
        waktu_jawab = answer.get('waktu_jawab')  # Waktu yang diambil untuk menjawab soal
        
        # Dapatkan soal dari database berdasarkan id_soal
        soal = Soal.query.filter_by(id=id_soal, id_quiz_kode=id_quiz_kode, deleted_at=None).first()
        if not soal:
            continue

        # Dapatkan total waktu dari soal
        total_waktu = soal.waktu.detik  # Anggap `total_waktu` dalam satuan detik

        # Tentukan apakah jawaban benar atau salah
        is_correct = False
        correct_jawaban = None
        for j in soal.soal_jawaban:
            if j.text_jawaban == jawaban_text and j.correct:
                is_correct = True
                correct_jawaban = j  # Simpan jawaban yang benar untuk akses bobot

        score_for_question = 0
        if correct_jawaban and is_correct:
            # Hitung score untuk soal ini berdasarkan waktu jawab dan bobot jawaban yang benar
            score_for_question = ((waktu_jawab / total_waktu) * 100) * correct_jawaban.bobot
            total_score += score_for_question

        # Simpan jawaban peserta
        peserta_jawaban = PesertaJawaban(
            id_quiz_kode=id_quiz_kode,
            id_soal=id_soal,
            id_peserta=id_peserta,
            jawaban=jawaban_text,
            iscorrect=is_correct,
            bobot=correct_jawaban.bobot if is_correct else 0,
            created_at=datetime.utcnow()
        )
        db.session.add(peserta_jawaban)

    # Cek apakah nilai peserta sudah ada sebelumnya
    peserta_nilai = PesertaNilai.query.filter_by(id_quiz_kode=id_quiz_kode, id_peserta=id_peserta, deleted_at=None).first()
    if peserta_nilai:
        # Tambahkan nilai baru ke nilai yang sudah ada
        peserta_nilai.nilai += total_score
        peserta_nilai.updated_at = datetime.utcnow()
    else:
        # Tambah nilai baru jika belum ada
        peserta_nilai = PesertaNilai(
            id_quiz_kode=id_quiz_kode,
            id_peserta=id_peserta,
            nilai=total_score,
            created_at=datetime.utcnow()
        )
        db.session.add(peserta_nilai)

    db.session.commit()

    return jsonify({
        'message': 'Jawaban dan nilai peserta berhasil disimpan', 
        'data':{
            'total_score': total_score,
            'score_for_question': score_for_question
        }
    }), 200



@quiz_bp.route('/answer_count/<int:id_quiz_kode>/<int:id_soal>', methods=['GET'])
def answer_count(id_quiz_kode, id_soal):
    # Dapatkan semua jawaban peserta untuk soal tertentu dalam quiz,
    # dengan peserta yang tidak memiliki deleted_at
    peserta_jawaban = (
        db.session.query(PesertaJawaban)
        .join(Peserta, Peserta.id == PesertaJawaban.id_peserta)
        .filter(
            PesertaJawaban.id_quiz_kode == id_quiz_kode,
            PesertaJawaban.id_soal == id_soal,
            PesertaJawaban.deleted_at == None,
            Peserta.deleted_at == None  # Tambahkan filter peserta.deleted_at IS NULL
        )
        .all()
    )

    # Hitung jumlah pilihan untuk setiap opsi jawaban
    answer_count = {}
    for jawaban in peserta_jawaban:
        if jawaban.jawaban not in answer_count:
            answer_count[jawaban.jawaban] = 0
        answer_count[jawaban.jawaban] += 1


    # Kembalikan hasil dalam format JSON
    return jsonify(answer_count), 200

@quiz_bp.route('/top_scores/<int:id_quiz_code>', methods=['GET'])
def get_top_scores(id_quiz_code):
    # Mengambil 5 data peserta nilai teratas berdasarkan id_quiz_code
    top_scores = (
        db.session.query(PesertaNilai, Peserta)
        .join(Peserta, Peserta.id == PesertaNilai.id_peserta)
        .filter(
            PesertaNilai.id_quiz_kode == id_quiz_code,
            PesertaNilai.deleted_at == None,
            Peserta.deleted_at == None
        )
        .order_by(desc(PesertaNilai.nilai))
        .limit(5)
        .all()
    )
    
    # Format data untuk respons JSON
    top_scores_data = [
        {
            "name": peserta.kode_unik,
            "score": peserta_nilai.nilai
        }
        for peserta_nilai, peserta in top_scores
    ]

    return jsonify(top_scores_data), 200

from sqlalchemy import func

@quiz_bp.route('/ranking/<int:id_quiz_kode>/<int:id_peserta>', methods=['GET'])
def get_participant_ranking(id_quiz_kode, id_peserta):
    # Buat subquery untuk menghitung peringkat berdasarkan nilai dan waktu tercepat
    subquery = (
        db.session.query(
            PesertaNilai.id_peserta,
            PesertaNilai.nilai,
            PesertaNilai.created_at,
            func.rank()
            .over(
                order_by=[
                    PesertaNilai.nilai.desc(),       # Prioritaskan nilai tertinggi
                    PesertaNilai.created_at.asc()   # Jika nilai sama, pilih waktu tercepat
                ]
            ).label("rank")
        )
        .join(Peserta, Peserta.id == PesertaNilai.id_peserta)  # Join dengan tabel Peserta
        .filter(
            PesertaNilai.id_quiz_kode == id_quiz_kode,
            PesertaNilai.deleted_at.is_(None),  # Pastikan nilai belum dihapus
            Peserta.deleted_at.is_(None)       # Pastikan peserta belum dihapus
        )
        .subquery()
    )

    # Gabungkan subquery dengan tabel Peserta untuk mendapatkan kode_unik peserta
    peserta_nilai = (
        db.session.query(
            Peserta.kode_unik,
            subquery.c.nilai,
            subquery.c.rank,
            subquery.c.created_at  # Ambil waktu untuk validasi
        )
        .join(Peserta, Peserta.id == subquery.c.id_peserta)
        .filter(
            Peserta.deleted_at.is_(None),       # Pastikan Peserta belum dihapus
            subquery.c.id_peserta == id_peserta  # Cari peserta berdasarkan ID
        )
        .first()
    )

    # Jika peserta tidak ditemukan, kembalikan error
    if not peserta_nilai:
        return jsonify({'error': 'Peserta tidak ditemukan'}), 404

    # Struktur respons dengan rank, kode_unik, nilai, dan waktu
    response = {
        "rank": peserta_nilai.rank,
        "kode_unik": peserta_nilai.kode_unik,  # Nama peserta
        "points": peserta_nilai.nilai,
        "created_at": peserta_nilai.created_at.strftime('%Y-%m-%d %H:%M:%S')  # Format waktu
    }

    return jsonify(response), 200


@quiz_bp.route('/top_winners/<int:id_quiz_kode>', methods=['GET'])
def get_top_winners(id_quiz_kode):
    # Ambil tiga peserta dengan nilai tertinggi berdasarkan id_quiz_kode
    top_scores = (db.session.query(Peserta.kode_unik.label("name"), PesertaNilai.nilai.label("score"))
                  .join(PesertaNilai, Peserta.id == PesertaNilai.id_peserta)
                  .filter(PesertaNilai.id_quiz_kode == id_quiz_kode, Peserta.deleted_at.is_(None), PesertaNilai.deleted_at.is_(None))
                  .order_by(PesertaNilai.nilai.desc(), PesertaNilai.created_at.asc())
                  .limit(3)
                  .all())

    # Jika tidak ada peserta yang ditemukan
    if not top_scores:
        return jsonify({'error': 'No winners found for this quiz code'}), 404

    # Map posisi ke dalam struktur data yang diinginkan
    winners = []
    for index, (name, score) in enumerate(top_scores):
        position = index + 1
        winner_data = {
            "position": position,
            "name": name,
            "score": score,
            "bgColor": "bg-red-400" if position == 2 else "bg-yellow-500" if position == 1 else "bg-green-400",
            "color": "bg-red-400" if position == 2 else "bg-yellow-400" if position == 1 else "bg-green-400",
            "height": "h-40" if position == 2 else "h-64" if position == 1 else "h-48",
            "avatar": f"/src/img/mendali_{position}.png"
        }
        winners.append(winner_data)

    return jsonify(winners), 200

# ==========================================Login=======================================
# Endpoint untuk registrasi (untuk membuat akun baru)
@quiz_bp.route('/register', methods=['POST'])
def register():
    data = request.get_json()
    username = data['username']
    password = data['password']
    
    hashed_password = generate_password_hash(password, method='pbkdf2:sha256')
    new_user = Users(username=username, password=hashed_password)
    db.session.add(new_user)
    db.session.commit()
    
    return jsonify({"message": "User registered successfully!"}), 201

# Endpoint untuk login
@quiz_bp.route('/login', methods=['POST'])
def login():
    data = request.get_json()
    username = data['username']
    password = data['password']
    
    user = Users.query.filter_by(username=username).first()
    
    if user and check_password_hash(user.password, password):
        return jsonify({"message": "Login successful", "username": username}), 200
    else:
        return jsonify({"message": "Invalid username or password"}), 401

# ====================================SESI MAIN VR=========================================
@quiz_bp.route('/sesi/vr', methods=['POST'])
def create_sesi():
    try:
        data = request.get_json()
        new_sesi = SesiMainVR(
            no_sesi=data['no_sesi'],
            nama=data['nama'],
            skenario=data['skenario'],
            kota=data['kota'],
            lokasi=data['lokasi'],
            waktu_mulai=datetime.fromisoformat(data['waktu_mulai']) if data.get('waktu_mulai') else None,
            waktu_selesai=datetime.fromisoformat(data['waktu_selesai']) if data.get('waktu_selesai') else None,
        )
        db.session.add(new_sesi)
        db.session.commit()
        return jsonify({"message": "Sesi created successfully", "data": new_sesi.to_dict()}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@quiz_bp.route('/sesi/vr', methods=['GET'])
def get_all_sesi():
    sesi_list = SesiMainVR.query.filter_by(deleted_at=None).all()
    return jsonify([sesi.to_dict() for sesi in sesi_list]), 200

@quiz_bp.route('/sesi/vr/<int:id>', methods=['PUT'])
def update_sesi(id):
    try:
        sesi = SesiMainVR.query.get(id)
        if not sesi or sesi.deleted_at:
            return jsonify({"error": "Sesi not found"}), 404

        data = request.get_json()
        sesi.no_sesi = data.get('no_sesi', sesi.no_sesi)
        sesi.nama = data.get('nama', sesi.nama)
        sesi.skenario = data.get('skenario', sesi.skenario)
        sesi.kota = data.get('kota', sesi.kota)
        sesi.lokasi = data.get('lokasi', sesi.lokasi)
        sesi.waktu_mulai = datetime.fromisoformat(data['waktu_mulai']) if data.get('waktu_mulai') else sesi.waktu_mulai
        sesi.waktu_selesai = datetime.fromisoformat(data['waktu_selesai']) if data.get('waktu_selesai') else sesi.waktu_selesai

        db.session.commit()
        return jsonify({"message": "Sesi updated successfully", "data": sesi.to_dict()}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@quiz_bp.route('/sesi/vr/<int:id>', methods=['DELETE'])
def delete_sesi(id):
    try:
        sesi = SesiMainVR.query.get(id)
        if not sesi or sesi.deleted_at:
            return jsonify({"error": "Sesi not found"}), 404

        sesi.deleted_at = datetime.utcnow()
        db.session.commit()
        return jsonify({"message": "Sesi deleted successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@quiz_bp.route('/import_sesi_main_vr', methods=['POST'])
def import_sesi_main_vr():
    if 'file' not in request.files:
        return jsonify({'message': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No selected file'}), 400

    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        imported_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Mulai dari baris kedua (lewat header)
            no_sesi, nama, skenario, kota, lokasi, waktu_mulai, waktu_selesai = row

            waktu_mulai_dt = datetime.strptime(waktu_mulai, '%Y-%m-%d %H:%M:%S') if waktu_mulai else None
            waktu_selesai_dt = datetime.strptime(waktu_selesai, '%Y-%m-%d %H:%M:%S') if waktu_selesai else None

            sesi = SesiMainVR(
                no_sesi=no_sesi,
                nama=nama,
                skenario=skenario,
                kota=kota,
                lokasi=lokasi,
                waktu_mulai=waktu_mulai_dt,
                waktu_selesai=waktu_selesai_dt
            )
            db.session.add(sesi)
            imported_data.append(sesi.to_dict())

        db.session.commit()
        return jsonify({'message': 'Data imported successfully', 'data': imported_data}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'message': 'Error importing data', 'error': str(e)}), 500



# ====================================SESI MAIN AR=========================================
@quiz_bp.route('/sesi/ar', methods=['POST'])
def create_sesi_ar():
    try:
        data = request.get_json()
        new_sesi = SesiMainAR(
            no_sesi=data['no_sesi'],
            nama=data['nama'], 
            kota=data['kota'],
            lokasi=data['lokasi'],
            waktu_mulai=datetime.fromisoformat(data['waktu_mulai']) if data.get('waktu_mulai') else None,
            waktu_selesai=datetime.fromisoformat(data['waktu_selesai']) if data.get('waktu_selesai') else None,
        )
        db.session.add(new_sesi)
        db.session.commit()
        return jsonify({"message": "Sesi created successfully", "data": new_sesi.to_dict()}), 201
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@quiz_bp.route('/sesi/ar', methods=['GET'])
def get_all_sesi_ar():
    sesi_list = SesiMainAR.query.filter_by(deleted_at=None).all()
    return jsonify([sesi.to_dict() for sesi in sesi_list]), 200

@quiz_bp.route('/sesi/ar/<int:id>', methods=['PUT'])
def update_sesi_ar(id):
    try:
        sesi = SesiMainAR.query.get(id)
        if not sesi or sesi.deleted_at:
            return jsonify({"error": "Sesi not found"}), 404

        data = request.get_json()
        sesi.no_sesi = data.get('no_sesi', sesi.no_sesi)
        sesi.nama = data.get('nama', sesi.nama) 
        sesi.kota = data.get('kota', sesi.kota)
        sesi.lokasi = data.get('lokasi', sesi.lokasi)
        sesi.waktu_mulai = datetime.fromisoformat(data['waktu_mulai']) if data.get('waktu_mulai') else sesi.waktu_mulai
        sesi.waktu_selesai = datetime.fromisoformat(data['waktu_selesai']) if data.get('waktu_selesai') else sesi.waktu_selesai

        db.session.commit()
        return jsonify({"message": "Sesi updated successfully", "data": sesi.to_dict()}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 400

@quiz_bp.route('/sesi/ar/<int:id>', methods=['DELETE'])
def delete_sesi_ar(id):
    try:
        sesi = SesiMainAR.query.get(id)
        if not sesi or sesi.deleted_at:
            return jsonify({"error": "Sesi not found"}), 404

        sesi.deleted_at = datetime.utcnow()
        db.session.commit()
        return jsonify({"message": "Sesi deleted successfully"}), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 400

# Endpoint untuk mengimpor data dari Excel
@quiz_bp.route('/import_sesi_main_ar', methods=['POST'])
def import_sesi_main_ar():
    if 'file' not in request.files:
        return jsonify({'message': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'message': 'No selected file'}), 400

    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        imported_data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Mulai dari baris kedua (header dilewati)
            no_sesi, nama, kota, lokasi, waktu_mulai, waktu_selesai = row

            # Konversi waktu mulai dan waktu selesai ke format datetime
            waktu_mulai_dt = datetime.strptime(waktu_mulai, '%Y-%m-%d %H:%M:%S') if waktu_mulai else None
            waktu_selesai_dt = datetime.strptime(waktu_selesai, '%Y-%m-%d %H:%M:%S') if waktu_selesai else None

            # Buat objek SesiMainAR
            sesi = SesiMainAR(
                no_sesi=no_sesi,
                nama=nama,
                kota=kota,
                lokasi=lokasi,
                waktu_mulai=waktu_mulai_dt,
                waktu_selesai=waktu_selesai_dt
            )
            db.session.add(sesi)
            imported_data.append(sesi.to_dict())

        db.session.commit()
        return jsonify({'message': 'Data imported successfully', 'data': imported_data}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'message': 'Error importing data', 'error': str(e)}), 500


# ==============================================LOG SESSION================================================
# Endpoint untuk mengambil semua data
@quiz_bp.route('/log_session', methods=['GET'])
def get_all_logs():
    logs = LogSession.query.all()
    return jsonify([log.to_dict() for log in logs])

# Endpoint untuk mengambil data berdasarkan ID
@quiz_bp.route('/log_session/<int:id>', methods=['GET'])
def get_log_by_id(id):
    log = LogSession.query.get(id)
    if log:
        return jsonify(log.to_dict())
    return jsonify({'message': 'Log not found'}), 404

# Endpoint untuk membuat data baru
@quiz_bp.route('/log_session', methods=['POST'])
def create_log():
    data = request.json
    jumlah = data.get('jumlah')
    produk = data.get('produk')
    sumber = data.get('sumber')
    sumber_link = data.get('sumber_link')

    new_log = LogSession(
        jumlah=jumlah,
        produk=produk,
        sumber=sumber,
        sumber_link=sumber_link
    )
    db.session.add(new_log)
    db.session.commit()
    return jsonify({'message': 'Log created successfully', 'log': new_log.to_dict()}), 201

# Endpoint untuk mengupdate data berdasarkan ID
@quiz_bp.route('/log_session/<int:id>', methods=['PUT'])
def update_log(id):
    log = LogSession.query.get(id)
    if not log:
        return jsonify({'message': 'Log not found'}), 404

    data = request.json
    log.jumlah = data.get('jumlah', log.jumlah)
    log.produk = data.get('produk', log.produk)
    log.sumber = data.get('sumber', log.sumber)
    log.sumber_link = data.get('sumber_link', log.sumber_link)

    db.session.commit()
    return jsonify({'message': 'Log updated successfully', 'log': log.to_dict()})

# Endpoint untuk menghapus data berdasarkan ID
@quiz_bp.route('/log_session/<int:id>', methods=['DELETE'])
def delete_log(id):
    log = LogSession.query.get(id)
    if not log:
        return jsonify({'message': 'Log not found'}), 404

    db.session.delete(log)
    db.session.commit()
    return jsonify({'message': 'Log deleted successfully'})

# ============================LOG SESSION KOTA===============================================
# Create - Tambah Data
@quiz_bp.route('/logsessionkota', methods=['POST'])
def create_logsessionkota():
    data = request.json
    new_entry = LogSessionKota(
        jumlah=data['jumlah'],
        produk=data['produk'],
        kota=data['kota'],
        sumber=data.get('sumber'),
        sumber_link=data.get('sumber_link')
    )
    db.session.add(new_entry)
    db.session.commit()
    return jsonify(new_entry.to_dict()), 201

# Read - Ambil Semua Data
@quiz_bp.route('/logsessionkota', methods=['GET'])
def get_all_logsessionkota():
    all_entries = LogSessionKota.query.all()
    return jsonify([entry.to_dict() for entry in all_entries])

# Read - Ambil Data Berdasarkan ID
@quiz_bp.route('/logsessionkota/<int:id>', methods=['GET'])
def get_logsessionkota_by_id(id):
    entry = LogSessionKota.query.get_or_404(id)
    return jsonify(entry.to_dict())

# Update - Perbarui Data Berdasarkan ID
@quiz_bp.route('/logsessionkota/<int:id>', methods=['PUT'])
def update_logsessionkota(id):
    data = request.json
    entry = LogSessionKota.query.get_or_404(id)
    
    entry.jumlah = data.get('jumlah', entry.jumlah)
    entry.produk = data.get('produk', entry.produk)
    entry.kota = data.get('kota', entry.kota)
    entry.sumber = data.get('sumber', entry.sumber)
    entry.sumber_link = data.get('sumber_link', entry.sumber_link)
    
    db.session.commit()
    return jsonify(entry.to_dict())

# Delete - Hapus Data Berdasarkan ID
@quiz_bp.route('/logsessionkota/<int:id>', methods=['DELETE'])
def delete_logsessionkota(id):
    entry = LogSessionKota.query.get_or_404(id)
    db.session.delete(entry)
    db.session.commit()
    return jsonify({"message": f"Entry with id {id} has been deleted."})